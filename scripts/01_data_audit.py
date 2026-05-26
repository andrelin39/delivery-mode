from __future__ import annotations

import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_FILE = PROJECT_ROOT / "data" / "raw" / "rawdata_all.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "outputs"
LOG_DIR = PROJECT_ROOT / "logs"
TARGET_SHEET = "用這一個"
DEFINITION_SHEET = "變項定義"

ID_COLUMNS = ["PARTO_NO", "BABY_NO", "CHART", "ACC_NO"]
CATEGORY_COLUMNS_TO_CHECK = ["DELIVER_MODE", "PGADM", "EDUCATION", "MARRIAGE"]


def ensure_dirs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)


def read_workbook() -> dict[str, pd.DataFrame]:
    if not RAW_FILE.exists():
        raise FileNotFoundError(f"Raw data file not found: {RAW_FILE}")
    return pd.read_excel(RAW_FILE, sheet_name=None, engine="openpyxl")


def normalize_missing(value: Any) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def get_original_headers(sheet_name: str) -> list[Any]:
    workbook = load_workbook(RAW_FILE, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()


def sheet_overview(workbook: dict[str, pd.DataFrame]) -> pd.DataFrame:
    records = []
    for sheet_name, df in workbook.items():
        records.append(
            {
                "sheet_name": sheet_name,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": "; ".join(map(str, df.columns)),
            }
        )
    return pd.DataFrame(records)


def missing_report(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for column in df.columns:
        series = df[column]
        missing_count = int(series.map(normalize_missing).sum())
        records.append(
            {
                "column": column,
                "dtype": str(series.dtype),
                "rows": len(series),
                "missing_count": missing_count,
                "missing_pct": missing_count / len(series) if len(series) else np.nan,
                "non_missing_count": int(len(series) - missing_count),
                "unique_count_including_na": int(series.nunique(dropna=False)),
                "unique_count_excluding_na": int(series.nunique(dropna=True)),
            }
        )
    return pd.DataFrame(records).sort_values(
        ["missing_count", "column"], ascending=[False, True]
    )


def numeric_summary(df: pd.DataFrame) -> pd.DataFrame:
    numeric_df = df.select_dtypes(include="number")
    if numeric_df.empty:
        return pd.DataFrame()
    summary = numeric_df.describe(percentiles=[0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99]).T
    summary.insert(0, "column", summary.index)
    summary = summary.reset_index(drop=True)
    return summary


def category_value_counts(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for column in df.columns:
        series = df[column]
        is_category_like = (
            series.dtype == "object"
            or str(series.dtype).startswith("category")
            or series.dtype == "bool"
            or series.nunique(dropna=True) <= 30
        )
        if not is_category_like:
            continue
        counts = series.astype("string").fillna("<MISSING>").value_counts(dropna=False)
        for value, count in counts.items():
            records.append(
                {
                    "column": column,
                    "value": value,
                    "count": int(count),
                    "pct": count / len(series) if len(series) else np.nan,
                }
            )
    return pd.DataFrame(records)


def duplicate_column_names() -> pd.DataFrame:
    headers = get_original_headers(TARGET_SHEET)
    counts = Counter(headers)
    records = [
        {"column_name": column, "duplicate_count": count}
        for column, count in counts.items()
        if column is not None and count > 1
    ]
    return pd.DataFrame(records)


def add_issue(
    issues: list[dict[str, Any]],
    issue_type: str,
    column: str | None,
    count: int | None,
    details: str,
) -> None:
    issues.append(
        {
            "issue_type": issue_type,
            "column": column,
            "count": count,
            "details": details,
        }
    )


def existing_columns(df: pd.DataFrame, columns: list[str]) -> list[str]:
    return [column for column in columns if column in df.columns]


def numeric_series(df: pd.DataFrame, column: str) -> pd.Series:
    return pd.to_numeric(df[column], errors="coerce")


def flag_range(
    df: pd.DataFrame,
    issues: list[dict[str, Any]],
    column: str,
    low: float | None,
    high: float | None,
    issue_type: str,
) -> None:
    if column not in df.columns:
        add_issue(issues, "missing_expected_column", column, None, "Column is not present.")
        return
    series = numeric_series(df, column)
    mask = pd.Series(False, index=df.index)
    clauses = []
    if low is not None:
        mask = mask | (series < low)
        clauses.append(f"< {low}")
    if high is not None:
        mask = mask | (series > high)
        clauses.append(f"> {high}")
    count = int(mask.sum())
    if count:
        examples = series[mask].dropna().head(10).tolist()
        add_issue(
            issues,
            issue_type,
            column,
            count,
            f"Values outside {' or '.join(clauses)}. Examples: {examples}",
        )


def find_age_columns(df: pd.DataFrame) -> list[str]:
    exact = [column for column in ["AGE", "MATERNAL_AGE", "MOM_AGE"] if column in df.columns]
    if exact:
        return exact
    return [
        column
        for column in df.columns
        if "AGE" in str(column).upper()
        and "GEST" not in str(column).upper()
        and "STAGE" not in str(column).upper()
    ]


def extract_defined_values(definition_df: pd.DataFrame | None, variable: str) -> set[str]:
    if definition_df is None or definition_df.empty:
        return set()

    variable_upper = variable.upper()
    values: set[str] = set()
    pattern = re.compile(r"(^|[,\s;/])([A-Za-z0-9_.-]+)\s*[:=：]", re.UNICODE)

    for _, row in definition_df.iterrows():
        cells = ["" if pd.isna(value) else str(value).strip() for value in row.tolist()]
        row_text = " ".join(cells).upper()
        if variable_upper not in row_text:
            continue

        for cell in cells:
            for match in pattern.finditer(cell):
                values.add(match.group(2).strip())

        for cell in cells[1:]:
            if re.fullmatch(r"[A-Za-z0-9_.-]+", cell):
                values.add(cell)

    metadata_words = {
        "NOMINAL",
        "ORDINAL",
        "CONTINUOUS",
        "INTERVAL",
        "RATIO",
        "STRING",
        "NUMERIC",
        "DATE",
        "DATETIME",
    }
    return {
        value
        for value in values
        if value and value.upper() != variable_upper and value.upper() not in metadata_words
    }


def potential_data_issues(
    df: pd.DataFrame, definition_df: pd.DataFrame | None
) -> dict[str, pd.DataFrame]:
    issues: list[dict[str, Any]] = []

    duplicate_columns = duplicate_column_names()
    if not duplicate_columns.empty:
        add_issue(
            issues,
            "duplicate_column_names",
            None,
            int(duplicate_columns["duplicate_count"].sum()),
            "Original header row contains duplicated column names.",
        )

    duplicate_row_mask = df.duplicated(keep=False)
    duplicate_rows = df.loc[duplicate_row_mask].copy()
    if duplicate_row_mask.any():
        add_issue(
            issues,
            "duplicate_rows",
            None,
            int(duplicate_row_mask.sum()),
            "Fully duplicated rows found.",
        )

    duplicate_id_records = []
    for column in ID_COLUMNS:
        if column not in df.columns:
            add_issue(issues, "missing_expected_id_column", column, None, "Column is not present.")
            continue
        non_missing = df[column].dropna()
        duplicate_mask = non_missing.duplicated(keep=False)
        duplicate_count = int(duplicate_mask.sum())
        if duplicate_count:
            add_issue(
                issues,
                "duplicate_id_values",
                column,
                duplicate_count,
                "Duplicated non-missing ID values found.",
            )
            counts = non_missing[duplicate_mask].value_counts()
            for value, count in counts.items():
                duplicate_id_records.append(
                    {"column": column, "value": value, "duplicate_count": int(count)}
                )

    for column in find_age_columns(df):
        flag_range(df, issues, column, 10, 60, "age_out_of_expected_range")

    for column in [column for column in df.columns if "BMI" in str(column).upper()]:
        flag_range(df, issues, column, 10, 80, "bmi_out_of_expected_range")

    gest_columns = ["GESTAT_WEEK", "GESTAT_DAY", "GESTAT_DAYS-ALL"]
    missing_gest = [column for column in gest_columns if column not in df.columns]
    if missing_gest:
        for column in missing_gest:
            add_issue(issues, "missing_expected_gestation_column", column, None, "Column is not present.")
    else:
        weeks = numeric_series(df, "GESTAT_WEEK")
        days = numeric_series(df, "GESTAT_DAY")
        gestat_days_all = numeric_series(df, "GESTAT_DAYS-ALL")
        calculated_days = weeks * 7 + days
        calculated_decimal_weeks = weeks + (days / 7)
        appears_decimal_weeks = gestat_days_all.dropna().max() < 100

        if appears_decimal_weeks:
            mismatch = (
                gestat_days_all.notna()
                & calculated_decimal_weeks.notna()
                & ~np.isclose(gestat_days_all, calculated_decimal_weeks, atol=0.001)
            )
        else:
            mismatch = (
                gestat_days_all.notna()
                & calculated_days.notna()
                & (gestat_days_all != calculated_days)
            )

        if mismatch.any():
            examples = pd.DataFrame(
                {
                    "GESTAT_WEEK": weeks[mismatch],
                    "GESTAT_DAY": days[mismatch],
                    "GESTAT_DAYS-ALL": gestat_days_all[mismatch],
                    "calculated_days": calculated_days[mismatch],
                    "calculated_decimal_weeks": calculated_decimal_weeks[mismatch],
                }
            ).head(10)
            add_issue(
                issues,
                "gestational_age_inconsistency",
                "GESTAT_WEEK/GESTAT_DAY/GESTAT_DAYS-ALL",
                int(mismatch.sum()),
                "GESTAT_DAYS-ALL differs from calculated gestational age. "
                f"Interpretation used: {'decimal_weeks' if appears_decimal_weeks else 'total_days'}. "
                f"Examples: {examples.to_dict('records')}",
            )
        if appears_decimal_weeks:
            add_issue(
                issues,
                "gestat_days_all_definition_check",
                "GESTAT_DAYS-ALL",
                int(gestat_days_all.notna().sum()),
                "Values appear to be decimal gestational weeks, not total days; confirm variable definition.",
            )

    flag_range(df, issues, "STAGE1_MINS", 0, 1440, "stage1_minutes_negative_or_extreme")
    flag_range(df, issues, "STAGE2_MINS", 0, 240, "stage2_minutes_negative_or_extreme")
    flag_range(df, issues, "BIRTH_WEIGHT", 300, 6000, "birth_weight_out_of_expected_range")
    flag_range(df, issues, "BIRTH_LENGTH", 20, 70, "birth_length_out_of_expected_range")
    flag_range(df, issues, "A_S_1", 0, 10, "apgar_out_of_expected_range")
    flag_range(df, issues, "A_S_5", 0, 10, "apgar_out_of_expected_range")

    undefined_records = []
    for column in CATEGORY_COLUMNS_TO_CHECK:
        if column not in df.columns:
            add_issue(issues, "missing_expected_category_column", column, None, "Column is not present.")
            continue
        defined_values = extract_defined_values(definition_df, column)
        observed = df[column].dropna().astype("string").str.strip()
        observed = observed[observed != ""]
        if not defined_values or len(defined_values) < 2:
            add_issue(
                issues,
                "category_definition_needs_manual_review",
                column,
                int(observed.nunique()),
                "Could not reliably extract allowed values from variable definition sheet.",
            )
            for value, count in observed.value_counts().items():
                undefined_records.append(
                    {
                        "column": column,
                        "value": value,
                        "count": int(count),
                        "status": "needs_manual_definition_review",
                    }
                )
            continue
        undefined = observed[~observed.isin(defined_values)]
        if not undefined.empty:
            add_issue(
                issues,
                "category_value_not_in_definition",
                column,
                int(undefined.count()),
                f"Allowed values parsed from definition sheet: {sorted(defined_values)}",
            )
            for value, count in undefined.value_counts().items():
                undefined_records.append(
                    {
                        "column": column,
                        "value": value,
                        "count": int(count),
                        "status": "not_in_parsed_definition",
                    }
                )

    issues_summary = pd.DataFrame(issues)
    if issues_summary.empty:
        issues_summary = pd.DataFrame(columns=["issue_type", "column", "count", "details"])

    return {
        "issues_summary": issues_summary,
        "duplicate_column_names": duplicate_columns,
        "duplicate_id_values": pd.DataFrame(duplicate_id_records),
        "duplicate_rows": duplicate_rows,
        "category_definition_check": pd.DataFrame(undefined_records),
    }


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    try:
        writer = pd.ExcelWriter(path, engine="xlsxwriter")
    except ModuleNotFoundError:
        writer = pd.ExcelWriter(path, engine="openpyxl")

    with writer:
        for sheet_name, data in sheets.items():
            safe_sheet_name = sheet_name[:31]
            data.to_excel(writer, sheet_name=safe_sheet_name, index=False)


def main() -> None:
    ensure_dirs()
    workbook = read_workbook()
    if TARGET_SHEET not in workbook:
        raise ValueError(f"Target sheet '{TARGET_SHEET}' not found. Sheets: {list(workbook)}")

    target_df = workbook[TARGET_SHEET]
    definition_df = workbook.get(DEFINITION_SHEET)

    overview = sheet_overview(workbook)
    missing = missing_report(target_df)
    numeric = numeric_summary(target_df)
    categories = category_value_counts(target_df)
    issues = potential_data_issues(target_df, definition_df)

    write_excel(OUTPUT_DIR / "01_sheet_overview.xlsx", {"sheet_overview": overview})
    write_excel(
        OUTPUT_DIR / "02_missing_report.xlsx",
        {
            "missing_report": missing,
            "dtypes": pd.DataFrame(
                {"column": target_df.columns, "dtype": [str(dtype) for dtype in target_df.dtypes]}
            ),
        },
    )
    write_excel(OUTPUT_DIR / "03_numeric_summary.xlsx", {"numeric_summary": numeric})
    write_excel(OUTPUT_DIR / "04_category_value_counts.xlsx", {"value_counts": categories})
    write_excel(OUTPUT_DIR / "05_potential_data_issues.xlsx", issues)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_path = LOG_DIR / "01_data_audit.log"
    log_path.write_text(
        "\n".join(
            [
                f"Run timestamp: {timestamp}",
                f"Raw file: {RAW_FILE}",
                f"Target sheet: {TARGET_SHEET}",
                f"Rows: {len(target_df)}",
                f"Columns: {len(target_df.columns)}",
                f"Issue rows: {len(issues['issues_summary'])}",
            ]
        ),
        encoding="utf-8",
    )

    print("Data audit completed.")
    print(f"Target sheet rows: {len(target_df)}")
    print(f"Target sheet columns: {len(target_df.columns)}")
    print(f"Potential issue types: {issues['issues_summary']['issue_type'].nunique()}")
    print(f"Outputs written to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
